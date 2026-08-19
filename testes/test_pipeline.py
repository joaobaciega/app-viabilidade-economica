"""T16 — o build REPROVA snapshot invalido. Uma fixture por regra S1-S13.

Plano §6.4: "Alguem renomeia coluna e o app quebra -> validacao de schema na
publicacao, com ERRO CLARO em vez de tela branca."

Cada teste abaixo viola UMA regra e verifica que ela e a que reprova. Depois,
test_T16_publicar_reprova_com_exit_1 verifica que o processo de publicacao
inteiro devolve exit != 0 e nao escreve snapshot.
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import pytest

from pipeline import esquema
from pipeline.publicar import ErroDePlanilha, ler_planilha, main, proxima_versao
from pipeline.validacoes import validar

HOJE = date.today().isoformat()


def _preco_valido(**overrides) -> dict:
    linha = {
        "marca": "volkswagen",
        "modelo": "polo",
        "ano_ini": 2023,
        "ano_fim": 2025,
        "posicao": "motorista",
        "unidade": "par",
        "preco": 477.17,
        "url_print": "prints/volkswagen_polo_600_2026-03-10.png",
        "data_coleta": "10/03/2026",
        "tipo_fonte": "loja_oficial_ml",
        "tipo_par": "par_nativo",
        "print_sem_banner_conferido": True,
        "selo_oficial_conferido": True,
    }
    linha.update(overrides)
    return linha


def _aplicacao_para(linha: dict) -> dict:
    return {
        "marca": linha["marca"],
        "modelo": linha["modelo"],
        "ano_ini": linha["ano_ini"],
        "ano_fim": linha["ano_fim"],
        "posicao": linha["posicao"],
        "medida_mm": 600,
    }


def _tabelas(precos: list[dict], catalogo: list[dict] | None = None) -> dict:
    return {
        "modelos": [],
        "aplicacao": [_aplicacao_para(p) for p in precos],
        "precos_originais": precos,
        "catalogo_refil": catalogo if catalogo is not None else [],
    }


def _regras(falhas) -> set[str]:
    return {f.regra for f in falhas}


# ---------------------------------------------------------------------------
# O caso valido precisa passar, senao os testes negativos nao provam nada
# ---------------------------------------------------------------------------


def test_registro_valido_passa() -> None:
    assert validar(_tabelas([_preco_valido()])) == []


def test_catalogo_vazio_passa() -> None:
    """Catalogo sem registros e o estado ESPERADO desta fase (§7.3)."""
    assert validar(_tabelas([], catalogo=[])) == []


# ---------------------------------------------------------------------------
# S1-S13, uma fixture invalida por regra
# ---------------------------------------------------------------------------


def test_S1_unidade_ausente() -> None:
    falhas = validar(_tabelas([_preco_valido(unidade=None)]))
    assert "S1" in _regras(falhas)


def test_S1_unidade_fora_do_enum() -> None:
    falhas = validar(_tabelas([_preco_valido(unidade="pares")]))
    assert "S1" in _regras(falhas)


def test_S2_data_ausente_ou_ilegivel() -> None:
    assert "S2" in _regras(validar(_tabelas([_preco_valido(data_coleta=None)])))
    assert "S2" in _regras(validar(_tabelas([_preco_valido(data_coleta="ontem")])))


def test_S2_data_no_futuro() -> None:
    amanha = (date.today() + timedelta(days=1)).strftime("%d/%m/%Y")
    falhas = validar(_tabelas([_preco_valido(data_coleta=amanha)]))
    assert "S2" in _regras(falhas)


def test_S3_print_ausente() -> None:
    falhas = validar(_tabelas([_preco_valido(url_print=None)]))
    assert "S3" in _regras(falhas)


def test_S3_print_inexistente_no_disco(tmp_path: Path) -> None:
    falhas = validar(_tabelas([_preco_valido()]), raiz_prints=tmp_path)
    assert "S3" in _regras(falhas)


def test_S4_nome_do_print_fora_do_padrao() -> None:
    falhas = validar(_tabelas([_preco_valido(url_print="prints/polo.png")]))
    assert "S4" in _regras(falhas)


def test_S5_banner_nao_conferido() -> None:
    """O print nao pode conter "nao e compativel com seu veiculo"."""
    falhas = validar(_tabelas([_preco_valido(print_sem_banner_conferido=False)]))
    assert "S5" in _regras(falhas)


def test_S6_ano_fim_menor_que_ano_ini() -> None:
    falhas = validar(_tabelas([_preco_valido(ano_ini=2025, ano_fim=2023)]))
    assert "S6" in _regras(falhas)


def test_S6_faixas_de_ano_sobrepostas() -> None:
    a = _preco_valido(ano_ini=2020, ano_fim=2024)
    b = _preco_valido(ano_ini=2023, ano_fim=2026)
    falhas = validar(_tabelas([a, b]))
    assert "S6" in _regras(falhas)


def test_S7_par_composto_com_uma_parcela() -> None:
    """Um cartao com METADE do preco e pior que nenhum cartao."""
    linha = _preco_valido(tipo_par="par_composto", grupo_par_id="polo-2023")
    falhas = validar(_tabelas([linha]))
    assert "S7" in _regras(falhas)


def test_S7_par_composto_com_duas_parcelas_passa() -> None:
    a = _preco_valido(
        tipo_par="par_composto",
        grupo_par_id="polo-2023",
        posicao="motorista",
    )
    b = _preco_valido(
        tipo_par="par_composto",
        grupo_par_id="polo-2023",
        posicao="passageiro",
        url_print="prints/volkswagen_polo_400_2026-03-10.png",
    )
    assert "S7" not in _regras(validar(_tabelas([a, b])))


def test_S8_loja_oficial_sem_selo_conferido() -> None:
    """Sem o selo, o print prova preco, NAO PROVA ORIGEM."""
    falhas = validar(_tabelas([_preco_valido(selo_oficial_conferido=False)]))
    assert "S8" in _regras(falhas)


def test_S9_preco_zero_ou_ausente() -> None:
    assert "S9" in _regras(validar(_tabelas([_preco_valido(preco=0)])))
    assert "S9" in _regras(validar(_tabelas([_preco_valido(preco=None)])))


def test_S9_indisponivel_dispensa_preco_e_print() -> None:
    """Regra de fallback do plano §2.4: "sem preco oficial publicado"."""
    linha = _preco_valido(
        tipo_fonte="indisponivel", preco=None, url_print=None,
        print_sem_banner_conferido=False, selo_oficial_conferido=False,
    )
    falhas = validar(_tabelas([linha]))
    assert _regras(falhas) - {"S10"} == set(), falhas


def test_S10_sem_linha_de_aplicacao_com_medida() -> None:
    """Sem medida o cartao NAO E AUDITAVEL."""
    linha = _preco_valido()
    tabelas = {
        "modelos": [],
        "aplicacao": [],  # nenhuma medida
        "precos_originais": [linha],
        "catalogo_refil": [],
    }
    assert "S10" in _regras(validar(tabelas))


def test_S10_medida_nao_e_duplicada_no_schema() -> None:
    """`posicao` e `medida_mm` tem UMA fonte autoritativa: a aba `aplicacao`.

    Duplicar garante divergencia, e divergencia de medida num cartao que existe
    para ser auditavel e o pior lugar possivel para ela aparecer.
    """
    assert "medida_mm" in esquema.APLICACAO.nomes
    assert "medida_mm" not in esquema.PRECOS_ORIGINAIS.nomes


def test_S11_campo_de_sessao_no_snapshot() -> None:
    """O snapshot e PUBLICO. Dado de reuniao nunca sai da sessao."""
    linha = _preco_valido()
    linha["palhetas_originais_mes"] = 20
    falhas = validar(_tabelas([linha]))
    assert "S11" in _regras(falhas)


def test_S12_custo_da_suicatech_no_snapshot() -> None:
    """A tabela de preco da Suicatech nao vive no snapshot (plano §6.4)."""
    linha = _preco_valido()
    linha["custo_aquisicao"] = 84.90
    falhas = validar(_tabelas([linha]))
    assert "S12" in _regras(falhas)


def test_S13_catalogo_sem_dianteiro() -> None:
    """A Tela 1 abre em dianteiro."""
    catalogo = [
        {
            "sku": "TR-001",
            "categoria": "traseiro",
            "unidade": "unitario",
            "medida_min_mm": 250,
            "medida_max_mm": 400,
        }
    ]
    falhas = validar(_tabelas([], catalogo=catalogo))
    assert "S13" in _regras(falhas)


def test_S13_catalogo_com_dianteiro_passa() -> None:
    catalogo = [
        {
            "sku": "DI-001",
            "categoria": "dianteiro",
            "unidade": "par",
            "medida_min_mm": 350,
            "medida_max_mm": 700,
        }
    ]
    assert validar(_tabelas([], catalogo=catalogo)) == []


# ---------------------------------------------------------------------------
# T16 — o build reprova de ponta a ponta
# ---------------------------------------------------------------------------


def _planilha_com(tmp_path: Path, precos: list[dict]) -> Path:
    import openpyxl

    caminho = tmp_path / "planilha.xlsx"
    livro = openpyxl.Workbook()
    livro.remove(livro.active)

    for aba in esquema.ABAS:
        planilha = livro.create_sheet(aba.nome)
        planilha.append(list(aba.nomes))
        if aba.nome == "precos_originais":
            for linha in precos:
                planilha.append([linha.get(c) for c in aba.nomes])
        elif aba.nome == "aplicacao":
            for linha in precos:
                app = _aplicacao_para(linha)
                planilha.append([app.get(c) for c in aba.nomes])

    livro.save(caminho)
    return caminho


def test_T16_publicar_reprova_com_exit_1(tmp_path: Path) -> None:
    """A definicao de pronto: o build FALHA quando alguma regra reprova."""
    invalido = _preco_valido(unidade="pares", preco=-1)
    planilha = _planilha_com(tmp_path, [invalido])
    saida = tmp_path / "snapshot"

    codigo = main(
        ["--planilha", str(planilha), "--saida", str(saida)]
    )

    assert codigo == 1, "build precisa reprovar com exit != 0"
    assert not list(saida.glob("*.json")) if saida.exists() else True, (
        "NADA pode ser publicado quando uma validacao reprova"
    )


def test_T16_publicar_aceita_planilha_valida(tmp_path: Path) -> None:
    planilha = _planilha_com(tmp_path, [_preco_valido()])
    saida = tmp_path / "snapshot"

    codigo = main(["--planilha", str(planilha), "--saida", str(saida)])

    assert codigo == 0
    assert (saida / "snapshot_v1.json").exists()
    assert (saida / "ultimo.json").exists()


def test_T16_versao_do_snapshot_incrementa(tmp_path: Path) -> None:
    planilha = _planilha_com(tmp_path, [_preco_valido()])
    saida = tmp_path / "snapshot"

    main(["--planilha", str(planilha), "--saida", str(saida)])
    main(["--planilha", str(planilha), "--saida", str(saida)])

    assert (saida / "snapshot_v1.json").exists()
    assert (saida / "snapshot_v2.json").exists()
    assert proxima_versao(saida) == 3


def test_coluna_renomeada_da_erro_claro(tmp_path: Path) -> None:
    """Plano §6.4: erro claro em vez de tela branca."""
    import openpyxl

    caminho = tmp_path / "renomeada.xlsx"
    livro = openpyxl.Workbook()
    livro.remove(livro.active)
    for aba in esquema.ABAS:
        planilha = livro.create_sheet(aba.nome)
        nomes = list(aba.nomes)
        if aba.nome == "precos_originais":
            nomes[nomes.index("unidade")] = "unidade_de_venda"  # renomeada
        planilha.append(nomes)
    livro.save(caminho)

    with pytest.raises(ErroDePlanilha, match="unidade"):
        ler_planilha(caminho)


def test_aba_ausente_da_erro_claro(tmp_path: Path) -> None:
    import openpyxl

    caminho = tmp_path / "sem_aba.xlsx"
    livro = openpyxl.Workbook()
    livro.remove(livro.active)
    planilha = livro.create_sheet("modelos")
    planilha.append(list(esquema.MODELOS.nomes))
    livro.save(caminho)

    with pytest.raises(ErroDePlanilha, match="aplicacao"):
        ler_planilha(caminho)


def test_publicar_com_planilha_ausente_devolve_2(tmp_path: Path) -> None:
    codigo = main(["--planilha", str(tmp_path / "nao_existe.xlsx")])
    assert codigo == 2
