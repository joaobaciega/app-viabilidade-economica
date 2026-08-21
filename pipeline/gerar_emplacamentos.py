"""Publicacao da base de emplacamentos: xlsx -> JSON versionado.

Mesma fronteira de `publicar.py`: RODA NA SUA MAQUINA OU NO CI, NUNCA NO
NAVEGADOR. Este modulo nao e importado por src/ nem por app.py.

POR QUE ESTE PASSO EXISTE, e nao ler o xlsx direto na Tela 2: ler xlsx em
runtime exigiria openpyxl no requirements.txt, e o checklist PROIBE
(testes/test_checklist.py, test_dependencias_de_teste_nao_estao_no_runtime).
Cada pacote a menos e menos tempo entre o vendedor abrir o link e o cliente ver
a tela — o Community Cloud hiberna em 12 h e a chance de estar dormindo na
visita e alta (plano §9, risco 6). O JSON gerado aqui nao custa dependencia
nenhuma: a Tela 2 le com o `json` da biblioteca padrao.

Uso:
    python -m pipeline.gerar_emplacamentos
    python -m pipeline.gerar_emplacamentos --conferir      # valida sem gravar
    python -m pipeline.gerar_emplacamentos --base outra.xlsx

Codigo de saida:
    0  gerado (ou conferencia sem falhas)
    2  erro de leitura da planilha (aba ausente, coluna renomeada, base vazia)

NENHUM VALOR E ESTIMADO. Celula vazia vira `null` e a tela escreve "não
publicado". A propria base declara isso nas notas de metodo: "Nenhum valor foi
estimado, interpolado ou completado."
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
BASE_PADRAO = RAIZ / "emplacamentos_brasil_base.xlsx"
SAIDA_PADRAO = RAIZ / "dados" / "emplacamentos.json"

ABA = "base"
SCHEMA_VERSAO = 1

# As colunas do xlsx que precisam existir. Renomear uma delas para o build
# aqui, com mensagem legivel, em vez de sair um JSON pela metade.
COLUNAS = (
    "marca",
    "modelo",
    "categoria",
    "posicao_na_marca",
    "emplacamentos_2026_ytd",
    "emplacamentos_2025_ytd",
    "emplacamentos_2025",
    "variacao_pct",
    "criterio",
    "fonte",
    "url",
    "data_consulta",
    "observacao",
)

# Rotulos das tres janelas. Ficam AQUI e nao na tela: sao um atributo do
# recorte da fonte, e mudam junto com a edicao do informativo.
JANELAS = {
    "atual": "jan–jul/2026",
    "anterior": "jan–jul/2025",
    "fechado": "ano 2025",
}

# A frase de cobertura que a base repete em TODA linha da marca. O gerador
# separa esse sufixo do que e especifico do modelo, para a tela mostrar a
# cobertura uma vez por marca em vez de cinco vezes iguais.
MARCA_COBERTURA = "Cobertura da marca na fonte:"


class ErroDeBase(Exception):
    """Aba ausente, coluna renomeada ou base vazia. Erro claro, nao tela branca."""


# ---------------------------------------------------------------------------
# Leitura
# ---------------------------------------------------------------------------


def ler_base(caminho: Path) -> list[dict]:
    """Le a aba `base` e devolve uma lista de dicionarios, uma por modelo."""
    import openpyxl

    if not caminho.exists():
        raise ErroDeBase(f"base não encontrada: {caminho}")

    livro = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    try:
        if ABA not in livro.sheetnames:
            raise ErroDeBase(
                f"aba obrigatória '{ABA}' ausente. "
                f"Abas encontradas: {livro.sheetnames}"
            )

        linhas = livro[ABA].iter_rows(values_only=True)
        try:
            cabecalho = [str(c).strip() if c is not None else "" for c in next(linhas)]
        except StopIteration:
            raise ErroDeBase(f"aba '{ABA}' está completamente vazia")

        faltando = [c for c in COLUNAS if c not in cabecalho]
        if faltando:
            raise ErroDeBase(
                f"aba '{ABA}': coluna(s) ausente(s) ou renomeada(s): {faltando}. "
                f"Cabeçalho encontrado: {cabecalho}"
            )

        registros = [
            {k: _normalizar(v) for k, v in zip(cabecalho, valores) if k}
            for valores in linhas
            if valores is not None and any(v is not None for v in valores)
        ]
    finally:
        livro.close()

    if not registros:
        raise ErroDeBase(f"aba '{ABA}' não tem nenhuma linha de dado")

    return registros


def _normalizar(valor: object) -> object:
    """Converte tipos do openpyxl em algo serializavel em JSON."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, str):
        return valor.strip()
    return valor


# ---------------------------------------------------------------------------
# Montagem
# ---------------------------------------------------------------------------


def _unico(registros: list[dict], coluna: str) -> str:
    """O valor de uma coluna que a base mantem igual nas 80 linhas.

    Sobe para o topo do JSON em vez de repetir por modelo. Se um dia a base
    passar a ter mais de um valor — outra edicao do informativo, outra data de
    coleta — isso e uma MUDANCA DE RECORTE, e o build para aqui para que a
    decisao seja tomada de proposito e nao por acidente.
    """
    valores = {str(r[coluna]) for r in registros if r.get(coluna)}
    if len(valores) != 1:
        raise ErroDeBase(
            f"coluna '{coluna}' precisa ter um único valor em toda a base, "
            f"e tem {len(valores)}: {sorted(valores)[:3]}"
        )
    return valores.pop()


def _partir_observacao(texto: str) -> tuple[str, str]:
    """Separa a nota do MODELO da frase de cobertura da MARCA.

    A base concatena as duas no mesmo campo: o que e especifico do modelo vem
    primeiro, e a frase "Cobertura da marca na fonte: ..." vem no fim, igual em
    todas as linhas da marca.
    """
    if not texto:
        return "", ""
    corte = texto.find(MARCA_COBERTURA)
    if corte < 0:
        return texto.strip(), ""
    return texto[:corte].strip(), texto[corte + len(MARCA_COBERTURA) :].strip()


def montar(registros: list[dict]) -> dict:
    """Agrupa por marca, na ordem em que a base lista, e monta o JSON."""
    marcas: dict[str, dict] = {}

    for linha, r in enumerate(registros, start=2):  # +2: cabecalho e base 1
        nome = str(r["marca"])
        nota_modelo, cobertura = _partir_observacao(str(r.get("observacao") or ""))

        marca = marcas.setdefault(nome, {"cobertura": "", "modelos": []})
        if cobertura and not marca["cobertura"]:
            marca["cobertura"] = cobertura

        # A posicao ORDENA a lista. Celula vazia ou texto aqui derrubaria o
        # sort() com um TypeError cru; o build para antes, com o numero da
        # linha da planilha, que e o que a curadoria precisa para corrigir.
        try:
            posicao = int(r["posicao_na_marca"])
        except (TypeError, ValueError):
            raise ErroDeBase(
                f"linha {linha} ({nome} {r.get('modelo')}): "
                f"'posicao_na_marca' precisa ser um número inteiro, e é "
                f"{r['posicao_na_marca']!r}"
            )

        marca["modelos"].append(
            {
                "modelo": r["modelo"],
                "categoria": r["categoria"],
                "posicao": posicao,
                "atual": r["emplacamentos_2026_ytd"],
                "anterior": r["emplacamentos_2025_ytd"],
                "fechado": r["emplacamentos_2025"],
                "variacao": r["variacao_pct"],
                "nota": nota_modelo,
            }
        )

    for marca in marcas.values():
        marca["modelos"].sort(key=lambda m: m["posicao"])

    return {
        "schema_versao": SCHEMA_VERSAO,
        "criterio": _unico(registros, "criterio"),
        "fonte": _unico(registros, "fonte"),
        "url": _unico(registros, "url"),
        "data_consulta": _unico(registros, "data_consulta"),
        "janelas": JANELAS,
        "marcas": dict(sorted(marcas.items())),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m pipeline.gerar_emplacamentos",
        description="Converte a base de emplacamentos em JSON para a Tela 2.",
    )
    parser.add_argument("--base", type=Path, default=BASE_PADRAO)
    parser.add_argument("--saida", type=Path, default=SAIDA_PADRAO)
    parser.add_argument(
        "--conferir", action="store_true", help="valida e reporta, sem gravar"
    )
    args = parser.parse_args(argv)

    # flush=True: sem isso o stdout com buffer aparece DEPOIS do stderr e o
    # relatorio fica ilegivel no terminal.
    print(f"Lendo {args.base}", flush=True)
    try:
        registros = ler_base(args.base)
        dados = montar(registros)
    except ErroDeBase as erro:
        print(f"\nERRO DE BASE: {erro}", file=sys.stderr, flush=True)
        return 2

    modelos = sum(len(m["modelos"]) for m in dados["marcas"].values())
    print(f"  {len(dados['marcas'])} marca(s), {modelos} modelo(s)", flush=True)
    print(f"  fonte coletada em {dados['data_consulta']}", flush=True)

    sem_2025 = [
        f"{marca} {m['modelo']}"
        for marca, dados_marca in dados["marcas"].items()
        for m in dados_marca["modelos"]
        if m["fechado"] is None
    ]
    if sem_2025:
        # NAO e falha: a base declara que celula vazia significa numero nao
        # encontrado na fonte. O relatorio existe para a curadoria conferir.
        print(
            f"  {len(sem_2025)} modelo(s) sem número de {JANELAS['fechado']} "
            f"(a tela escreve 'não publicado'): {', '.join(sem_2025)}",
            flush=True,
        )

    if args.conferir:
        print("\nModo --conferir: nada gravado.", flush=True)
        return 0

    args.saida.parent.mkdir(parents=True, exist_ok=True)
    args.saida.write_text(
        json.dumps(dados, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    # `relative_to` levanta se --saida apontar para fora do repo, e falhar
    # DEPOIS de gravar seria o pior dos dois mundos.
    try:
        destino = args.saida.resolve().relative_to(RAIZ)
    except ValueError:
        destino = args.saida
    print(f"\nGerado: {destino} (schema v{SCHEMA_VERSAO})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
