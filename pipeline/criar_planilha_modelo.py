"""Gera dados/planilha_modelo.xlsx — abas com cabecalho correto e ZERO registros.

O escopo desta entrega e explicito: nenhum dado real e responsabilidade desta
fase. A planilha nasce VAZIA, com os cabecalhos que o schema exige, para que:

  1. o pipeline tenha o que validar de ponta a ponta
  2. voce tenha o formato exato para comecar a Fase 0
  3. as Telas 2 e 3 exibam o estado vazio em vez de numero inventado

NUNCA semeie esta planilha com numeros plausiveis. O plano §2.4 e literal:
"uma linha inventada destroi as outras 200".

Uso:  python -m pipeline.criar_planilha_modelo
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from pipeline import esquema

RAIZ = Path(__file__).resolve().parents[1]
DESTINO = RAIZ / "dados" / "planilha_modelo.xlsx"

_AJUDA = {
    "modelos": "Emplacamentos por marca e modelo. Fonte: informes PÚBLICOS da Fenabrave (PDF), nunca área logada.",
    "aplicacao": "Fonte AUTORITATIVA de posicao e medida. precos_originais referencia esta aba (S10).",
    "precos_originais": "Preco da palheta original. unidade por linha (par/unitario). Print com selo oficial e sem o banner de incompatibilidade.",
    "catalogo_refil": "Catalogo do refil. dianteiro = par, traseiro = unitario. Nunca derive um do outro.",
}


def criar(destino: Path = DESTINO) -> Path:
    livro = openpyxl.Workbook()
    livro.remove(livro.active)

    cabecalho_fonte = Font(bold=True, size=11)
    cabecalho_fundo = PatternFill("solid", fgColor="F5F5F5")

    for aba in esquema.ABAS:
        planilha = livro.create_sheet(aba.nome)

        for coluna, definicao in enumerate(aba.colunas, start=1):
            celula = planilha.cell(row=1, column=coluna, value=definicao.nome)
            celula.font = cabecalho_fonte
            celula.fill = cabecalho_fundo
            celula.alignment = Alignment(vertical="center")

            largura = max(14, min(30, len(definicao.nome) + 4))
            planilha.column_dimensions[celula.column_letter].width = largura

            # A ajuda por coluna vive num COMENTARIO de celula, nao numa linha
            # de dado. Uma linha de ajuda na planilha e lida pelo pipeline como
            # registro e reprova o build — o que ja aconteceu uma vez aqui.
            nota = ["obrigatória" if definicao.obrigatoria else "opcional"]
            nota.append(f"tipo: {definicao.tipo}")
            if definicao.enum:
                nota.append("valores: " + " | ".join(definicao.enum))
            celula.comment = Comment("\n".join(nota), "schema")

        # Os dados comecam na linha 2, imediatamente sob o cabecalho.
        planilha.freeze_panes = "A2"

    # Uma aba de instrucoes, para quem abrir a planilha em seis meses.
    guia = livro.create_sheet("_LEIA", 0)
    guia["A1"] = "Planilha de curadoria — refil de palhetas (Suicatech)"
    guia["A1"].font = Font(bold=True, size=14)
    linhas = [
        "",
        "Cada aba tem só o cabeçalho, na linha 1. Comece a preencher na linha 2.",
        "A descrição de cada coluna está no COMENTÁRIO da célula do cabeçalho",
        "(passe o mouse sobre o nome da coluna).",
        "",
        "Publicação:  python -m pipeline.publicar",
        "Conferir sem publicar:  python -m pipeline.publicar --conferir",
        "",
        "O build REPROVA se qualquer validação S1–S13 falhar, e nada é publicado.",
        "Isso é deliberado: um erro claro aqui vale mais que uma tela quebrada",
        "na frente do cliente.",
        "",
        "Duas regras que nunca podem ser quebradas:",
        "  1. NUNCA preencher preço de vendedor terceiro e chamar de 'original'.",
        "     Uma linha inventada destrói as outras 200.",
        "  2. NUNCA colocar custo de aquisição ou preço de venda da Suicatech",
        "     nesta planilha. O snapshot é público (S11 e S12 reprovam o build).",
        "",
        "Abas:",
    ]
    for i, texto in enumerate(linhas, start=2):
        guia[f"A{i}"] = texto
    proxima = len(linhas) + 2
    for aba in esquema.ABAS:
        guia[f"A{proxima}"] = f"  {aba.nome}: {_AJUDA.get(aba.nome, '')}"
        proxima += 1
    guia.column_dimensions["A"].width = 100

    destino.parent.mkdir(parents=True, exist_ok=True)
    livro.save(destino)
    return destino


if __name__ == "__main__":
    caminho = criar()
    print(f"Criada: {caminho}")
    print("Abas:", ", ".join(a.nome for a in esquema.ABAS))
    print("Zero registros — nenhum dado real é responsabilidade desta fase.")
