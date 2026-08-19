"""Publicacao: planilha -> validacao de schema -> snapshot JSON versionado.

Plano §6.4: "Planilha alimentando o app funciona SE o schema for rigido e SE
houver um passo de publicacao. Sem esse passo intermediario, 'voce edita e o app
le' carrega vazio. Com ele, voce publica quando quiser."

RODA NA SUA MAQUINA OU NO CI, NUNCA NO NAVEGADOR. Este modulo nao e importado
por src/ nem por app.py (fronteira verificada por AST em test_checklist.py).

Uso:
    python -m pipeline.publicar
    python -m pipeline.publicar --planilha dados/planilha_modelo.xlsx
    python -m pipeline.publicar --conferir     # valida sem publicar

Codigo de saida:
    0  publicado (ou conferencia sem falhas)
    1  QUALQUER validacao S1-S13 reprovou — build reprovado, nada e publicado
    2  erro de leitura da planilha (aba ausente, coluna renomeada)

`git` nao e necessario: o versionamento e por nome de arquivo e pelo campo
`versao_snapshot`. Ter git ajuda voce a voltar atras, mas o pipeline nao depende.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from pipeline import esquema
from pipeline.validacoes import Falha, validar

RAIZ = Path(__file__).resolve().parents[1]
PLANILHA_PADRAO = RAIZ / "dados" / "planilha_modelo.xlsx"
SAIDA_PADRAO = RAIZ / "dados" / "snapshot"
PRINTS_PADRAO = RAIZ / "dados" / "prints"

PADRAO_VERSAO = re.compile(r"^snapshot_v(\d+)\.json$")


class ErroDePlanilha(Exception):
    """Aba ausente ou coluna renomeada. Erro CLARO em vez de tela branca."""


# ---------------------------------------------------------------------------
# Leitura
# ---------------------------------------------------------------------------


def ler_planilha(caminho: Path) -> dict[str, list[dict]]:
    """Le as abas declaradas em esquema.ABAS e devolve listas de dicionarios.

    Levanta ErroDePlanilha com mensagem legivel se uma aba obrigatoria faltar
    ou se uma coluna obrigatoria tiver sido renomeada — que e exatamente o
    risco que o plano §6.4 aponta.
    """
    import openpyxl

    if not caminho.exists():
        raise ErroDePlanilha(f"planilha não encontrada: {caminho}")

    livro = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    tabelas: dict[str, list[dict]] = {}

    for aba in esquema.ABAS:
        if aba.nome not in livro.sheetnames:
            if aba.obrigatoria:
                raise ErroDePlanilha(
                    f"aba obrigatória '{aba.nome}' ausente. "
                    f"Abas encontradas: {livro.sheetnames}"
                )
            tabelas[aba.nome] = []
            continue

        planilha = livro[aba.nome]
        linhas = planilha.iter_rows(values_only=True)
        try:
            cabecalho = [
                str(c).strip() if c is not None else "" for c in next(linhas)
            ]
        except StopIteration:
            raise ErroDePlanilha(f"aba '{aba.nome}' está completamente vazia")

        faltando = [
            c.nome
            for c in aba.colunas
            if c.obrigatoria and c.nome not in cabecalho
        ]
        if faltando:
            raise ErroDePlanilha(
                f"aba '{aba.nome}': coluna(s) obrigatória(s) ausente(s) ou "
                f"renomeada(s): {faltando}. Cabeçalho encontrado: {cabecalho}"
            )

        registros: list[dict] = []
        for valores in linhas:
            if valores is None or all(v is None for v in valores):
                continue  # linha em branco
            registros.append(dict(zip(cabecalho, valores)))

        tabelas[aba.nome] = registros

    livro.close()
    return tabelas


# ---------------------------------------------------------------------------
# Normalizacao
# ---------------------------------------------------------------------------


def _normalizar(valor: object) -> object:
    """Converte tipos do openpyxl em algo serializavel em JSON."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, str):
        return valor.strip()
    return valor


def _normalizar_tabelas(tabelas: dict[str, list[dict]]) -> dict[str, list[dict]]:
    return {
        aba: [
            {k: _normalizar(v) for k, v in registro.items() if k}
            for registro in registros
        ]
        for aba, registros in tabelas.items()
    }


# ---------------------------------------------------------------------------
# Versionamento
# ---------------------------------------------------------------------------


def proxima_versao(saida: Path) -> int:
    versoes = [
        int(m.group(1))
        for arquivo in saida.glob("snapshot_v*.json")
        if (m := PADRAO_VERSAO.match(arquivo.name))
    ]
    return max(versoes, default=0) + 1


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m pipeline.publicar",
        description="Valida a planilha e publica um snapshot JSON versionado.",
    )
    parser.add_argument("--planilha", type=Path, default=PLANILHA_PADRAO)
    parser.add_argument("--saida", type=Path, default=SAIDA_PADRAO)
    parser.add_argument("--prints", type=Path, default=PRINTS_PADRAO)
    parser.add_argument(
        "--conferir",
        action="store_true",
        help="valida e reporta, sem publicar",
    )
    args = parser.parse_args(argv)

    # flush=True em tudo: sem isso o stdout com buffer aparece DEPOIS do stderr
    # e o relatorio fica ilegivel no terminal.
    print(f"Lendo {args.planilha}", flush=True)
    try:
        tabelas = _normalizar_tabelas(ler_planilha(args.planilha))
    except ErroDePlanilha as erro:
        print(f"\nERRO DE PLANILHA: {erro}", file=sys.stderr, flush=True)
        return 2

    for aba, registros in tabelas.items():
        print(f"  {aba}: {len(registros)} registro(s)", flush=True)

    raiz_prints = args.prints if args.prints.exists() else None
    falhas: list[Falha] = validar(tabelas, raiz_prints=raiz_prints)

    if falhas:
        sys.stdout.flush()
        print(
            f"\n{len(falhas)} falha(s) de validação. NADA foi publicado.\n",
            file=sys.stderr,
            flush=True,
        )
        for falha in falhas:
            print(f"  {falha}", file=sys.stderr, flush=True)
        print(
            "\nCorrija a planilha e rode de novo. O build reprova de propósito: "
            "um erro claro aqui vale mais que uma tela quebrada na frente do "
            "cliente.",
            file=sys.stderr,
            flush=True,
        )
        return 1

    print("\nValidações S1–S13: todas passaram.", flush=True)

    if args.conferir:
        print("Modo --conferir: nada publicado.", flush=True)
        return 0

    args.saida.mkdir(parents=True, exist_ok=True)
    versao = proxima_versao(args.saida)
    snapshot = {
        "versao_snapshot": versao,
        "publicado_em": date.today().strftime("%d/%m/%Y"),
        "schema_versao": esquema.SCHEMA_VERSAO,
        **tabelas,
    }

    destino = args.saida / f"snapshot_v{versao}.json"
    texto = json.dumps(snapshot, ensure_ascii=False, indent=2)
    destino.write_text(texto, encoding="utf-8")
    (args.saida / "ultimo.json").write_text(texto, encoding="utf-8")

    print(f"Publicado: {destino.name} (schema v{esquema.SCHEMA_VERSAO})")
    print(f"Ponteiro atualizado: ultimo.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
